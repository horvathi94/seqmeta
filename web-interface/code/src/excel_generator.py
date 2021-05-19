import os
from datetime import datetime
from openpyxl import Workbook


class ExcelGenerator:

    save_dir = "/app/temp";

    def __init__(self):
        pass;


    @staticmethod
    def create_worksheet(title):
        wb = Workbook();
        ws = wb.active;
        ws.title = title;
        return wb, ws;


    @classmethod
    def save_excel(cls, wb, filename):
        excel_file = os.path.join(cls.save_dir, filename);
        wb.save(filename=excel_file);


    @classmethod
    def populate_gisaid(cls, ws, samples):
        for indx, sample in enumerate(samples):
            i = indx+3;
            ws["A{:d}".format(i)] = "Submitter"          # Submitter
            ws["B{:d}".format(i)] = "filename";          # Filename
            ws["C{:d}".format(i)] = "virus-name";        # Virus name
            ws["D{:d}".format(i)] = "betacoronavirus"    # Leave as default
            ws["E{:d}".format(i)] = sample["passage_details"]
            ws["F{:d}".format(i)] = str(sample["collection_date"]);
            ws["G{:d}".format(i)] = sample["location"];
            ws["H{:d}".format(i)] = sample["additional_location_info"];
            ws["I{:d}".format(i)] = sample["host"];
            ws["J{:d}".format(i)] = sample["additional_host_info"];
            ws["K{:d}".format(i)] = sample["sampling_strategy"];
            ws["L{:d}".format(i)] = sample["patient_gender"]
            ws["M{:d}".format(i)] = sample["patient_age"];
            ws["N{:d}".format(i)] = sample["patient_status"];
            ws["O{:d}".format(i)] = sample["specimen_source"];
            ws["P{:d}".format(i)] = sample["outbreak"];
            ws["Q{:d}".format(i)] = sample["last_vaccinated"];
            ws["R{:d}".format(i)] = sample["treatment"];
            ws["S{:d}".format(i)] = sample["sequencing_technology"];
            ws["T{:d}".format(i)] = sample["assembly_method"];
            ws["U{:d}".format(i)] = sample["coverage"];
            ws["V{:d}".format(i)] = sample["originating_lab_name"];
            ws["W{:d}".format(i)] = sample["originating_lab_address"];
            ws["X{:d}".format(i)] = sample["sample_name"];
            ws["Y{:d}".format(i)] = sample["submitting_lab_name"];
            ws["Z{:d}".format(i)] = sample["submitting_lab_address"];
            ws["AA{:d}".format(i)] = sample["sample_name"];
            ws["AB{:d}".format(i)] = sample["authors_list"];


    @classmethod
    def write_gisaid(cls, samples, filename=""):
        if filename == "":
            filename = str(datetime.now()) + "_gisaid";
        wb, ws = cls.create_worksheet("Submissions");
        cls.populate_gisaid(ws, samples);
        cls.save_excel(wb, filename+".xls");


